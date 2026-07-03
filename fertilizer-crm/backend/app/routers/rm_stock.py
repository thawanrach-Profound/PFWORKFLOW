from datetime import date
from decimal import Decimal
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.crm import RmStockSession, RmStockEntry, PurchaseOrder, RmSale, RmEntryTypeEnum, PoStatusEnum
from app.schemas.crm import (
    RmSessionCreate, RmSessionOut, RmStockEntryOut,
    PoCreate, PoOut,
    RmSaleCreate, RmSaleOut,
    ProductionUsageIn,
)

router = APIRouter(prefix="/api/rm-stock", tags=["rm-stock"])


def _calc_totals(s: RmStockSession):
    s.total_rm_value = sum(e.total_value or 0 for e in s.entries if e.entry_type == RmEntryTypeEnum.RAW)
    s.total_fg_value = sum(e.total_value or 0 for e in s.entries if e.entry_type == RmEntryTypeEnum.FG)
    s.total_paid_po  = sum(p.amount or 0 for p in s.pos if p.po_status == PoStatusEnum.PAID)
    s.total_pending_po = sum(p.amount or 0 for p in s.pos if p.po_status == PoStatusEnum.PENDING)


# ── Sessions ──────────────────────────────────────────────────
@router.get("", response_model=list[RmSessionOut])
def list_sessions(
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    db: Session = Depends(get_db),
):
    q = db.query(RmStockSession)
    if from_date: q = q.filter(RmStockSession.stock_date >= from_date)
    if to_date:   q = q.filter(RmStockSession.stock_date <= to_date)
    return q.order_by(RmStockSession.stock_date.desc()).all()


@router.get("/{session_id}", response_model=RmSessionOut)
def get_session(session_id: int, db: Session = Depends(get_db)):
    s = db.get(RmStockSession, session_id)
    if not s: raise HTTPException(404, "ไม่พบรอบ snapshot")
    return s


@router.post("", response_model=RmSessionOut, status_code=201)
def create_session(payload: RmSessionCreate, db: Session = Depends(get_db)):
    s = RmStockSession(
        stock_date=payload.stock_date,
        notes=payload.notes,
        created_by=payload.created_by,
    )
    db.add(s); db.flush()

    all_entries = (
        [(e, RmEntryTypeEnum.RAW) for e in payload.rm_entries] +
        [(e, RmEntryTypeEnum.FG)  for e in payload.fg_entries]
    )
    for e, etype in all_entries:
        entry = RmStockEntry(
            session_id=s.session_id,
            entry_type=etype,
            material_name=e.material_name,
            stock_qty_ton=e.stock_qty_ton,
            price_per_ton=e.price_per_ton,
            po_ref=e.po_ref,
            notes=e.notes,
            sort_order=e.sort_order,
        )
        db.add(entry)

    db.flush()
    _calc_totals(s)
    db.commit(); db.refresh(s)
    return s


@router.delete("/{session_id}", status_code=204)
def delete_session(session_id: int, db: Session = Depends(get_db)):
    s = db.get(RmStockSession, session_id)
    if not s: raise HTTPException(404, "ไม่พบรอบ snapshot")
    db.delete(s); db.commit()


# ── Production Usage (ตัดสต็อกจากรายงานเครื่องจักร) ──────────
@router.post("/production-usage", response_model=RmSessionOut, status_code=201)
def create_production_usage(payload: ProductionUsageIn, db: Session = Depends(get_db)):
    machines = ", ".join(payload.machine_nos) if payload.machine_nos else "ทุกเครื่อง"
    notes = payload.notes or f"ตัดจากรายงานการผลิต [{machines}]"
    s = RmStockSession(
        stock_date=payload.usage_date,
        notes=notes,
        created_by="machine-import",
    )
    db.add(s); db.flush()
    for m in payload.materials:
        if m.qty_kg <= 0:
            continue
        qty_ton = Decimal(str(round(m.qty_kg / 1000, 4)))
        db.add(RmStockEntry(
            session_id=s.session_id,
            entry_type=RmEntryTypeEnum.RAW,
            material_name=m.name,
            stock_qty_ton=-qty_ton,
            notes="ใช้ผลิต",
        ))
    db.flush()
    _calc_totals(s)
    db.commit(); db.refresh(s)
    return s


# ── Purchase Orders ───────────────────────────────────────────
@router.get("/po/list", response_model=list[PoOut])
def list_pos(
    po_status: Optional[str] = None,
    supplier: Optional[str] = None,
    session_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(PurchaseOrder)
    if po_status: q = q.filter(PurchaseOrder.po_status == po_status)
    if supplier:  q = q.filter(PurchaseOrder.supplier_name.ilike(f"%{supplier}%"))
    if session_id: q = q.filter(PurchaseOrder.session_id == session_id)
    return q.order_by(PurchaseOrder.created_at.desc()).all()


@router.post("/po", response_model=PoOut, status_code=201)
def create_po(payload: PoCreate, db: Session = Depends(get_db)):
    po = PurchaseOrder(**payload.model_dump())
    db.add(po); db.commit(); db.refresh(po)
    return po


@router.put("/po/{po_id}", response_model=PoOut)
def update_po(po_id: int, payload: PoCreate, db: Session = Depends(get_db)):
    po = db.get(PurchaseOrder, po_id)
    if not po: raise HTTPException(404, "ไม่พบ PO")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(po, k, v)
    db.commit(); db.refresh(po)
    return po


@router.delete("/po/{po_id}", status_code=204)
def delete_po(po_id: int, db: Session = Depends(get_db)):
    po = db.get(PurchaseOrder, po_id)
    if not po: raise HTTPException(404, "ไม่พบ PO")
    db.delete(po); db.commit()


# ── RM Sales (ขายวัตถุดิบออก) ────────────────────────────────
@router.get("/rm-sales", response_model=list[RmSaleOut])
def list_rm_sales(
    session_id: Optional[int] = None,
    buyer: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(RmSale)
    if session_id: q = q.filter(RmSale.session_id == session_id)
    if buyer: q = q.filter(RmSale.buyer_name.ilike(f"%{buyer}%"))
    return q.order_by(RmSale.sale_date.desc()).all()


@router.post("/rm-sales", response_model=RmSaleOut, status_code=201)
def create_rm_sale(payload: RmSaleCreate, db: Session = Depends(get_db)):
    sale = RmSale(**payload.model_dump())
    db.add(sale); db.commit(); db.refresh(sale)
    return sale


@router.put("/rm-sales/{rm_sale_id}", response_model=RmSaleOut)
def update_rm_sale(rm_sale_id: int, payload: RmSaleCreate, db: Session = Depends(get_db)):
    sale = db.get(RmSale, rm_sale_id)
    if not sale: raise HTTPException(404, "ไม่พบรายการ")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(sale, k, v)
    db.commit(); db.refresh(sale)
    return sale


@router.delete("/rm-sales/{rm_sale_id}", status_code=204)
def delete_rm_sale(rm_sale_id: int, db: Session = Depends(get_db)):
    sale = db.get(RmSale, rm_sale_id)
    if not sale: raise HTTPException(404, "ไม่พบรายการ")
    db.delete(sale); db.commit()


# ── Import Excel ──────────────────────────────────────────────

RAW_NAMES = {
    "ยูเรีย","แด็บ","8-40-0","ม็อบ","แอม","ดิน","OX",
    "18-46-0","21-0-0","20-20-0","46-0-0","0-0-60",
}

def _safe_float(v):
    try:
        if v is None: return None
        if isinstance(v, (int, float)): return float(v)
        return float(str(v).replace(",",""))
    except Exception: return None

def _safe_date(v):
    if v is None: return None
    if hasattr(v, "date"): return v.date()
    return None

def _parse_sheet_date(sheet_name: str):
    """แปลงชื่อ sheet เป็น date — รองรับ '16-1-69' และ 'วัตถูดิบ 26-6-69'"""
    import re
    m = re.search(r"(\d{1,2})[-/\s](\d{1,2})[-/\s](\d{2,4})", sheet_name)
    if not m: return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100: y += 2500
    try: return date(y - 543, mo, d)
    except Exception: return None


def _import_format_daily(ws_rows, s, db):
    """Format: สต๊อควัตถุดิบ+FG69.xlsx — ชื่อ Sheet = วันที่ dd-m-yy"""
    # PO จ่ายแล้ว (rows 4-30)
    for row in ws_rows[4:30]:
        if not row[1] or not row[3]: continue
        try:
            db.add(PurchaseOrder(
                session_id=s.session_id,
                supplier_name=str(row[1]).strip(),
                po_number=str(row[2]).strip() if row[2] else None,
                material_name=str(row[3]).strip(),
                qty_ton=_safe_float(row[4]),
                amount=_safe_float(row[5]),
                payment_date=_safe_date(row[6]),
                conditions=str(row[7]).strip() if row[7] else None,
                po_status=PoStatusEnum.PAID,
            ))
        except Exception: continue

    # PO pending (rows 34-43)
    for row in ws_rows[34:43]:
        if not row[1] or not row[3]: continue
        try:
            db.add(PurchaseOrder(
                session_id=s.session_id,
                supplier_name=str(row[1]).strip(),
                po_number=str(row[2]).strip() if row[2] else None,
                material_name=str(row[3]).strip(),
                qty_ton=_safe_float(row[4]),
                amount=_safe_float(row[5]),
                conditions=str(row[7]).strip() if row[7] else None,
                po_status=PoStatusEnum.PENDING,
            ))
        except Exception: continue

    # สต็อก วัตถุดิบ+FG (rows 52-104, col C=idx2)
    for idx, row in enumerate(ws_rows[52:104]):
        if not row[2]: continue
        name = str(row[2]).strip()
        if name in ("ยอดรวม", ""): continue
        try:
            qty = _safe_float(row[3])
            price = _safe_float(row[4])
            etype = RmEntryTypeEnum.RAW if name in RAW_NAMES else RmEntryTypeEnum.FG
            db.add(RmStockEntry(
                session_id=s.session_id,
                entry_type=etype,
                material_name=name,
                stock_qty_ton=qty,
                price_per_ton=price,
                sort_order=idx,
            ))
        except Exception: continue


def _import_format_snapshot(ws, ws_rows, stock_date, db):
    """Format: ตัวอยากข้อมูลเกี่ยวข้องวัตถุดิบ.xlsx — 2 sheets แยก"""
    sheet_name = ws.title

    if "การซื้อ" in sheet_name or "PO" in sheet_name.upper():
        # Sheet การซื้อวัตถุดิบ: rows 4+ = PO list
        # col: A=seq, B=supplier, C=PO#, D=material, E=qty, F=amount, G=payment_date, H=conditions, I=notes
        s = None
        for row in ws_rows[3:]:
            if not row[1] or not row[3]: continue
            if str(row[3]).strip() in ("สินค้า", "รวม", ""): continue
            if s is None:
                # หาวันที่จาก cell G2
                sd = _safe_date(ws_rows[1][6]) if len(ws_rows) > 1 else None
                if sd is None: sd = stock_date or date.today()
                s = RmStockSession(stock_date=sd, notes="import การซื้อวัตถุดิบ", created_by="import")
                db.add(s); db.flush()
            try:
                # ค่าขน = บางแถวมี supplier="ค่าขนส่ง" พ่วงท้าย
                is_freight = str(row[1]).strip() == "ค่าขนส่ง"
                supplier = str(row[1]).strip()
                material = str(row[3]).strip()
                qty = _safe_float(row[4])
                amount = _safe_float(row[5])
                payment_date = None
                g = row[6]
                if g and isinstance(g, str) and any(c.isdigit() for c in g):
                    # วันที่เป็น text เช่น "26 มิย.69"
                    payment_date = None  # parse ไม่ได้ reliably → เก็บใน notes
                elif hasattr(g, "date"):
                    payment_date = g.date()
                pay_note = str(g).strip() if g and not hasattr(g, "date") else None
                cond = str(row[7]).strip() if row[7] else None
                note_text = str(row[8]).strip() if row[8] else None
                if pay_note: note_text = (note_text or "") + f" วันจ่าย: {pay_note}"

                if is_freight:
                    # เพิ่ม freight_cost เข้า PO ล่าสุด
                    last_po = db.query(PurchaseOrder).filter(
                        PurchaseOrder.session_id == s.session_id
                    ).order_by(PurchaseOrder.po_id.desc()).first()
                    if last_po and amount:
                        last_po.freight_cost = amount
                    continue

                db.add(PurchaseOrder(
                    session_id=s.session_id,
                    supplier_name=supplier,
                    po_number=str(row[2]).strip() if row[2] else None,
                    material_name=material,
                    qty_ton=qty if isinstance(row[4], (int, float)) else None,
                    amount=amount,
                    payment_date=payment_date,
                    conditions=cond,
                    notes=note_text,
                    po_status=PoStatusEnum.PAID,
                ))
            except Exception: continue
        return s

    else:
        # Sheet วัตถูดิบ 26-6-69: ส่วนบน=ขายออก, ส่วนล่าง=สต็อก
        sd = stock_date or date.today()
        s = RmStockSession(stock_date=sd, notes=f"import {sheet_name}", created_by="import")
        db.add(s); db.flush()

        # ส่วนบน R1-R6: ขายวัตถุดิบออก (col A=seq, B=buyer, C=price, D=material, E=qty, F=amount, G=pay_date, H=cond)
        for row in ws_rows[1:6]:
            if not row[1] or not row[3]: continue
            if str(row[0] or "").strip() in ("ลำดับ", ""): continue
            try:
                db.add(RmSale(
                    session_id=s.session_id,
                    sale_date=sd,
                    buyer_name=str(row[1]).strip(),
                    material_name=str(row[3]).strip(),
                    qty_ton=_safe_float(row[4]),
                    price_per_ton=_safe_float(row[2]),
                    amount=_safe_float(row[5]),
                    payment_date=_safe_date(row[6]),
                    conditions=str(row[7]).strip() if row[7] else None,
                ))
            except Exception: continue

        # ส่วนล่าง R9+: สต็อก RAW (R10-R16) และ FG (R17+)
        for idx, row in enumerate(ws_rows[9:]):
            if not row[2]: continue
            name = str(row[2]).strip()
            if name in ("ยอดรวม", "สินค้า", "ลำดับ", ""): continue
            try:
                qty = _safe_float(row[3])
                price = _safe_float(row[4])
                # R10-R16 (idx 0-6) = RAW วัตถุดิบ
                etype = RmEntryTypeEnum.RAW if name in RAW_NAMES else RmEntryTypeEnum.FG
                db.add(RmStockEntry(
                    session_id=s.session_id,
                    entry_type=etype,
                    material_name=name,
                    stock_qty_ton=qty,
                    price_per_ton=price,
                    sort_order=idx,
                ))
            except Exception: continue
        return s


@router.post("/import", response_model=list[RmSessionOut], status_code=201)
async def import_rm_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import ไฟล์ Excel สต็อกวัตถุดิบ — รองรับ 2 format:
    1. สต๊อควัตถุดิบ+FG69.xlsx (ชื่อ Sheet=วันที่ dd-m-yy)
    2. ตัวอยากข้อมูลเกี่ยวข้องวัตถุดิบ.xlsx (Sheet 'วัตถูดิบ xx-x-xx' + 'การซื้อวัตถุดิบ')
    """
    from io import BytesIO
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    created_sessions = []

    # ตรวจ format: ถ้าชื่อ sheet มีคำว่า "วัตถุดิบ" หรือ "การซื้อ" = format ใหม่
    is_new_format = any(
        "วัตถุดิบ" in sh or "วัตถูดิบ" in sh or "การซื้อ" in sh
        for sh in wb.sheetnames
    )

    # หาวันที่รวมจากทุก sheet (สำหรับเชื่อม 2 sheet ให้เป็น session เดียว)
    snapshot_date = None
    for sh in wb.sheetnames:
        d = _parse_sheet_date(sh)
        if d: snapshot_date = d; break

    if is_new_format:
        # Format ใหม่: 2 sheets อาจสร้าง 2 session แต่ share วันที่
        created_session_ids = set()
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws_rows = list(ws.iter_rows(values_only=True))
            sd = _parse_sheet_date(sheet_name) or snapshot_date or date.today()
            s = _import_format_snapshot(ws, ws_rows, sd, db)
            if s is not None and s.session_id not in created_session_ids:
                db.flush()
                _calc_totals(s)
                created_sessions.append(s)
                created_session_ids.add(s.session_id)
    else:
        # Format เดิม: ชื่อ sheet = วันที่ dd-m-yy
        for sheet_name in wb.sheetnames:
            stock_date = _parse_sheet_date(sheet_name)
            if not stock_date: continue
            ws = wb[sheet_name]
            ws_rows = list(ws.iter_rows(values_only=True))
            note_date = None
            if len(ws_rows) > 1 and ws_rows[1][6] and hasattr(ws_rows[1][6], "year"):
                note_date = ws_rows[1][6].strftime("%d/%m/%Y")
            s = RmStockSession(
                stock_date=stock_date,
                notes=f"เช้าก่อนผลิต {note_date or sheet_name}",
                created_by="import",
            )
            db.add(s); db.flush()
            _import_format_daily(ws_rows, s, db)
            db.flush()
            _calc_totals(s)
            created_sessions.append(s)
    for s in created_sessions:
        db.refresh(s)
    return created_sessions
